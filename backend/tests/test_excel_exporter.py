"""Unit tests for the ExcelExporter service (ADR-0013, option B)."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import ColumnSchema
from app.services.excel_exporter import ExcelExporter


def _schema() -> ColumnSchema:
    return ColumnSchema(
        columns=[
            {"index": 1, "name": "Equipo_Local", "data_type": "texto", "example_value": "Borregos"},
            {"index": 2, "name": "Puntos_local", "data_type": "número", "example_value": "3"},
            {"index": 3, "name": "Estadio", "data_type": "texto", "example_value": "Disney"},
        ]
    )


def _sheet_from_bytes(content: bytes):
    return load_workbook(BytesIO(content)).active


@pytest.fixture
def exporter():
    return ExcelExporter()


def test_first_row_is_column_names(exporter):
    """Row 1 must be exactly the column names, in schema order."""
    content = exporter.build(_schema(), [])
    sheet = _sheet_from_bytes(content)

    assert [c.value for c in sheet[1]] == ["Equipo_Local", "Puntos_local", "Estadio"]


def test_no_scaffolding_rows_from_template(exporter):
    """The Tipo_Dato / Ejemplo_Valor template rows must NOT appear in the export."""
    records = [{"Equipo_Local": "Pumas", "Puntos_local": "4", "Estadio": "Azteca"}]
    content = exporter.build(_schema(), records)
    sheet = _sheet_from_bytes(content)

    # Header on row 1, data immediately on row 2 — no "texto"/"número" type row,
    # no "Borregos"/"Disney" example row in between.
    assert sheet.max_row == 2
    assert [c.value for c in sheet[2]] == ["Pumas", "4", "Azteca"]


def test_data_rows_follow_schema_order(exporter):
    """Each record is written in schema column order regardless of dict order."""
    records = [
        {"Estadio": "Azteca", "Equipo_Local": "Pumas", "Puntos_local": "4"},
        {"Puntos_local": "0", "Estadio": "BBVA", "Equipo_Local": "Rayados"},
    ]
    sheet = _sheet_from_bytes(exporter.build(_schema(), records))

    assert [c.value for c in sheet[2]] == ["Pumas", "4", "Azteca"]
    assert [c.value for c in sheet[3]] == ["Rayados", "0", "BBVA"]
    assert sheet.max_row == 3


def test_empty_records_yields_header_only(exporter):
    """With no records the workbook has just the header row."""
    sheet = _sheet_from_bytes(exporter.build(_schema(), []))
    assert sheet.max_row == 1


def test_missing_or_null_value_becomes_empty_cell(exporter):
    """A column absent from a record (or null) is written as an empty string."""
    records = [{"Equipo_Local": "Pumas", "Estadio": None}]  # Puntos_local missing
    sheet = _sheet_from_bytes(exporter.build(_schema(), records))

    # openpyxl may round-trip an empty string back as None; treat both as empty.
    assert [(c.value or "") for c in sheet[2]] == ["Pumas", "", ""]
