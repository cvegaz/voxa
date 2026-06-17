"""Excel file validation service.

Validates uploaded Excel files for structural correctness:
- File extension must be .xlsx
- Maximum 8 columns allowed
- Row 1 must have at least one non-empty column name
- Row 2 must have non-empty Tipo_Dato for all named columns
- Row 3 must have non-empty Ejemplo_Valor for all named columns
"""

import os
from io import BytesIO

from fastapi import UploadFile
from openpyxl import load_workbook

from app.models import ValidationResult


class ExcelValidator:
    """Validates uploaded Excel template files for structural correctness."""

    MAX_COLUMNS = 8
    ALLOWED_EXTENSIONS = {".xlsx"}

    def validate(self, file: UploadFile) -> ValidationResult:
        """Validate extension, structure, and content of the uploaded file.

        Validation is fail-fast: rejects at the first error found.

        Args:
            file: The uploaded file from FastAPI.

        Returns:
            ValidationResult with is_valid=True on success, or
            is_valid=False with error_code and detail on failure.
        """
        # 1. Validate file extension
        extension_result = self._validate_extension(file.filename)
        if not extension_result.is_valid:
            return extension_result

        # 2. Load workbook from file content
        try:
            file.file.seek(0)
            content = file.file.read()
            file.file.seek(0)
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception:
            return ValidationResult(
                is_valid=False,
                error_code="UNREADABLE_FILE",
                detail="El archivo está corrupto o no se puede leer como Excel.",
            )

        try:
            sheet = workbook.active

            # 3. Validate column count (1-8)
            column_result = self._validate_column_count(sheet)
            if not column_result.is_valid:
                return column_result

            # 4. Validate row 1 has at least one non-empty column name
            header_result = self._validate_header_row(sheet)
            if not header_result.is_valid:
                return header_result

            # Get the named columns (non-empty cells in row 1)
            named_columns = self._get_named_columns(sheet)

            # 5. Validate row 2 has non-empty Tipo_Dato for all named columns
            types_result = self._validate_data_types_row(sheet, named_columns)
            if not types_result.is_valid:
                return types_result

            # 6. Validate row 3 has non-empty Ejemplo_Valor for all named columns
            examples_result = self._validate_examples_row(sheet, named_columns)
            if not examples_result.is_valid:
                return examples_result

        finally:
            workbook.close()

        return ValidationResult(is_valid=True)

    def _validate_extension(self, filename: str | None) -> ValidationResult:
        """Check that the file has a .xlsx extension."""
        if filename is None:
            return ValidationResult(
                is_valid=False,
                error_code="INVALID_EXTENSION",
                detail="El archivo no tiene nombre. Solo se permiten archivos .xlsx.",
            )

        _, ext = os.path.splitext(filename)
        if ext.lower() not in self.ALLOWED_EXTENSIONS:
            return ValidationResult(
                is_valid=False,
                error_code="INVALID_EXTENSION",
                detail=f"El formato '{ext}' no es compatible. Solo se permiten archivos .xlsx.",
            )

        return ValidationResult(is_valid=True)

    def _validate_column_count(self, sheet) -> ValidationResult:
        """Validate the sheet has between 1 and MAX_COLUMNS columns with content."""
        # Count columns that have any content in the first 3 rows
        max_col = sheet.max_column
        if max_col is None or max_col == 0:
            return ValidationResult(
                is_valid=False,
                error_code="EMPTY_HEADER_ROW",
                detail="No se encontraron nombres de columna en la primera fila.",
            )

        if max_col > self.MAX_COLUMNS:
            return ValidationResult(
                is_valid=False,
                error_code="TOO_MANY_COLUMNS",
                detail=f"El archivo tiene {max_col} columnas. El límite permitido es {self.MAX_COLUMNS} columnas.",
            )

        return ValidationResult(is_valid=True)

    def _validate_header_row(self, sheet) -> ValidationResult:
        """Validate row 1 has at least one non-empty column name."""
        row1_values = []
        for cell in sheet[1]:
            value = cell.value
            if value is not None and str(value).strip() != "":
                row1_values.append(value)

        if not row1_values:
            return ValidationResult(
                is_valid=False,
                error_code="EMPTY_HEADER_ROW",
                detail="No se encontraron nombres de columna en la primera fila.",
            )

        return ValidationResult(is_valid=True)

    def _get_named_columns(self, sheet) -> list[tuple[int, str]]:
        """Get list of (column_index, name) for non-empty cells in row 1."""
        named = []
        for cell in sheet[1]:
            value = cell.value
            if value is not None and str(value).strip() != "":
                named.append((cell.column, str(value).strip()))
        return named

    def _validate_data_types_row(
        self, sheet, named_columns: list[tuple[int, str]]
    ) -> ValidationResult:
        """Validate row 2 has non-empty Tipo_Dato for all named columns."""
        missing = []
        for col_idx, col_name in named_columns:
            cell_value = sheet.cell(row=2, column=col_idx).value
            if cell_value is None or str(cell_value).strip() == "":
                missing.append(col_name)

        if missing:
            return ValidationResult(
                is_valid=False,
                error_code="MISSING_DATA_TYPES",
                detail=f"Faltan los tipos de dato en la fila 2 para las columnas: {', '.join(missing)}.",
                affected_columns=missing,
            )

        return ValidationResult(is_valid=True)

    def _validate_examples_row(
        self, sheet, named_columns: list[tuple[int, str]]
    ) -> ValidationResult:
        """Validate row 3 has non-empty Ejemplo_Valor for all named columns."""
        missing = []
        for col_idx, col_name in named_columns:
            cell_value = sheet.cell(row=3, column=col_idx).value
            if cell_value is None or str(cell_value).strip() == "":
                missing.append(col_name)

        if missing:
            return ValidationResult(
                is_valid=False,
                error_code="MISSING_EXAMPLES",
                detail=f"Faltan los ejemplos de valor en la fila 3 para las columnas: {', '.join(missing)}.",
                affected_columns=missing,
            )

        return ValidationResult(is_valid=True)
