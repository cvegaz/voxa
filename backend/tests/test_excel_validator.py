"""Unit tests for ExcelValidator service."""

import io
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook

from app.services.excel_validator import ExcelValidator


def _create_upload_file(filename: str, workbook: Workbook | None = None) -> MagicMock:
    """Helper to create a mock UploadFile with an in-memory Excel workbook."""
    mock_file = MagicMock()
    mock_file.filename = filename

    if workbook is not None:
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        mock_file.file = buffer
    else:
        mock_file.file = io.BytesIO(b"not an excel file")

    return mock_file


def _create_valid_workbook(num_columns: int = 3) -> Workbook:
    """Create a valid workbook with the given number of columns."""
    wb = Workbook()
    ws = wb.active
    for col in range(1, num_columns + 1):
        ws.cell(row=1, column=col, value=f"Columna_{col}")
        ws.cell(row=2, column=col, value=f"texto")
        ws.cell(row=3, column=col, value=f"ejemplo_{col}")
    return wb


class TestExcelValidatorExtension:
    """Tests for file extension validation."""

    def setup_method(self):
        self.validator = ExcelValidator()

    def test_valid_xlsx_extension(self):
        wb = _create_valid_workbook()
        file = _create_upload_file("data.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is True

    def test_invalid_csv_extension(self):
        file = _create_upload_file("data.csv")
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "INVALID_EXTENSION"

    def test_invalid_xls_extension(self):
        file = _create_upload_file("data.xls")
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "INVALID_EXTENSION"

    def test_no_extension(self):
        file = _create_upload_file("datafile")
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "INVALID_EXTENSION"

    def test_none_filename(self):
        file = _create_upload_file(None)
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "INVALID_EXTENSION"

    def test_uppercase_xlsx_extension(self):
        wb = _create_valid_workbook()
        file = _create_upload_file("DATA.XLSX", wb)
        result = self.validator.validate(file)
        assert result.is_valid is True


class TestExcelValidatorColumnCount:
    """Tests for column count validation."""

    def setup_method(self):
        self.validator = ExcelValidator()

    def test_one_column_valid(self):
        wb = _create_valid_workbook(1)
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is True

    def test_eight_columns_valid(self):
        wb = _create_valid_workbook(8)
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is True

    def test_nine_columns_rejected(self):
        wb = _create_valid_workbook(9)
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "TOO_MANY_COLUMNS"

    def test_twenty_columns_rejected(self):
        wb = _create_valid_workbook(20)
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "TOO_MANY_COLUMNS"


class TestExcelValidatorHeaderRow:
    """Tests for row 1 (header) validation."""

    def setup_method(self):
        self.validator = ExcelValidator()

    def test_empty_header_row_rejected(self):
        wb = Workbook()
        ws = wb.active
        # Row 1 is empty, but add content in row 2 to avoid empty sheet
        ws.cell(row=2, column=1, value="texto")
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "EMPTY_HEADER_ROW"

    def test_whitespace_only_header_rejected(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="   ")
        ws.cell(row=2, column=1, value="texto")
        ws.cell(row=3, column=1, value="ejemplo")
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "EMPTY_HEADER_ROW"


class TestExcelValidatorDataTypes:
    """Tests for row 2 (data types) validation."""

    def setup_method(self):
        self.validator = ExcelValidator()

    def test_missing_data_type_rejected(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Nombre")
        ws.cell(row=1, column=2, value="Edad")
        ws.cell(row=2, column=1, value="texto")
        # Column 2 row 2 is empty
        ws.cell(row=3, column=1, value="Juan")
        ws.cell(row=3, column=2, value="25")
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "MISSING_DATA_TYPES"
        assert "Edad" in result.affected_columns

    def test_all_data_types_present_valid(self):
        wb = _create_valid_workbook(3)
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is True


class TestExcelValidatorExamples:
    """Tests for row 3 (examples) validation."""

    def setup_method(self):
        self.validator = ExcelValidator()

    def test_missing_example_rejected(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Nombre")
        ws.cell(row=1, column=2, value="Edad")
        ws.cell(row=2, column=1, value="texto")
        ws.cell(row=2, column=2, value="número entero")
        ws.cell(row=3, column=1, value="Juan")
        # Column 2 row 3 is empty
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is False
        assert result.error_code == "MISSING_EXAMPLES"
        assert "Edad" in result.affected_columns

    def test_all_examples_present_valid(self):
        wb = _create_valid_workbook(4)
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is True


class TestExcelValidatorCorruptFile:
    """Tests for unreadable/corrupt file handling."""

    def setup_method(self):
        self.validator = ExcelValidator()

    def test_corrupt_file_rejected(self):
        mock_file = MagicMock()
        mock_file.filename = "corrupt.xlsx"
        mock_file.file = io.BytesIO(b"this is not a valid excel file at all")
        result = self.validator.validate(mock_file)
        assert result.is_valid is False
        assert result.error_code == "UNREADABLE_FILE"


class TestExcelValidatorSpecialCases:
    """Tests for special characters and edge cases."""

    def setup_method(self):
        self.validator = ExcelValidator()

    def test_columns_with_special_characters(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Dirección")
        ws.cell(row=1, column=2, value="Año de Nacimiento")
        ws.cell(row=2, column=1, value="texto")
        ws.cell(row=2, column=2, value="número entero")
        ws.cell(row=3, column=1, value="Calle Ñoño #123")
        ws.cell(row=3, column=2, value="1990")
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.is_valid is True

    def test_fail_fast_returns_first_error(self):
        """Validator should return TOO_MANY_COLUMNS before checking headers."""
        wb = _create_valid_workbook(10)
        file = _create_upload_file("test.xlsx", wb)
        result = self.validator.validate(file)
        assert result.error_code == "TOO_MANY_COLUMNS"
