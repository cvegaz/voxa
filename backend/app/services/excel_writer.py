"""Excel writer service.

Writes a DataFrame to an existing .xlsx file, preserving header rows (1-3)
and writing data starting from row 4. Uses openpyxl directly for control
over row placement.
"""

import pandas as pd
from openpyxl import load_workbook

from app.models import ColumnSchema


class ExcelWriter:
    """Writes DataFrame content to an existing Excel file, preserving headers."""

    def write(
        self,
        dataframe: pd.DataFrame,
        file_path: str,
        schema: ColumnSchema,
    ) -> None:
        """Export the DataFrame to the .xlsx file.

        Preserves header rows (1-3) and writes data starting from row 4.
        Overwrites the existing file on disk.

        Args:
            dataframe: DataFrame containing the data rows to write.
            file_path: Path to the existing .xlsx file.
            schema: ColumnSchema defining column order.

        Raises:
            FileNotFoundError: If the file at file_path does not exist.
        """
        try:
            workbook = load_workbook(file_path)
        except FileNotFoundError:
            raise FileNotFoundError(
                f"El archivo Excel no fue encontrado: {file_path}"
            )

        sheet = workbook.active

        # Clear existing data rows (row 4+)
        max_row = sheet.max_row
        if max_row >= 4:
            for row_idx in range(max_row, 3, -1):
                sheet.delete_rows(row_idx)

        # Write DataFrame rows starting at row 4
        column_names = [col.name for col in schema.columns]

        for row_offset, (_, row) in enumerate(dataframe.iterrows()):
            excel_row = 4 + row_offset
            for col_def in schema.columns:
                col_name = col_def.name
                col_idx = col_def.index
                value = row.get(col_name, "")
                if pd.isna(value):
                    value = ""
                sheet.cell(row=excel_row, column=col_idx, value=value)

        workbook.save(file_path)
