"""Unit tests for the DataFrameConverter service."""

import pandas as pd
import pytest
from openpyxl import Workbook

from app.models.template_models import ColumnDef, ColumnSchema
from app.services.dataframe_converter import DataFrameConverter


@pytest.fixture
def converter():
    """Return a DataFrameConverter instance."""
    return DataFrameConverter()


def _create_workbook(headers, types, examples, data_rows):
    """Helper to create a workbook with given header, metadata, and data rows.

    Args:
        headers: list of column name strings for row 1.
        types: list of data type strings for row 2.
        examples: list of example value strings for row 3.
        data_rows: list of lists, each inner list is a row of data (row 4+).
    """
    wb = Workbook()
    ws = wb.active

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for col_idx, dtype in enumerate(types, start=1):
        ws.cell(row=2, column=col_idx, value=dtype)

    for col_idx, example in enumerate(examples, start=1):
        ws.cell(row=3, column=col_idx, value=example)

    for row_offset, row_data in enumerate(data_rows, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_offset, column=col_idx, value=value)

    return wb


def _create_schema(headers, types, examples):
    """Helper to create a ColumnSchema from lists."""
    columns = [
        ColumnDef(index=i + 1, name=h, data_type=t, example_value=e)
        for i, (h, t, e) in enumerate(zip(headers, types, examples))
    ]
    return ColumnSchema(columns=columns)


class TestDataFrameConverterBasic:
    """Test basic conversion scenarios."""

    def test_single_column_single_row(self, converter):
        """Convert a workbook with one column and one data row."""
        wb = _create_workbook(
            headers=["Nombre"],
            types=["texto"],
            examples=["Juan"],
            data_rows=[["Carlos"]],
        )
        schema = _create_schema(["Nombre"], ["texto"], ["Juan"])

        df = converter.convert(wb, schema)

        assert isinstance(df, pd.DataFrame)
        assert list(df.columns) == ["Nombre"]
        assert len(df) == 1
        assert df.iloc[0]["Nombre"] == "Carlos"

    def test_multiple_columns_multiple_rows(self, converter):
        """Convert a workbook with multiple columns and rows."""
        headers = ["Nombre", "Edad", "Ciudad"]
        types = ["texto", "entero", "texto"]
        examples = ["Juan", "25", "Madrid"]
        data_rows = [
            ["Ana", 30, "Barcelona"],
            ["Luis", 28, "Valencia"],
            ["María", 35, "Sevilla"],
        ]

        wb = _create_workbook(headers, types, examples, data_rows)
        schema = _create_schema(headers, types, examples)

        df = converter.convert(wb, schema)

        assert list(df.columns) == ["Nombre", "Edad", "Ciudad"]
        assert len(df) == 3
        assert df.iloc[0]["Nombre"] == "Ana"
        assert df.iloc[1]["Edad"] == 28
        assert df.iloc[2]["Ciudad"] == "Sevilla"

    def test_empty_data_rows_returns_empty_dataframe(self, converter):
        """A workbook with no data rows (only header/metadata) returns empty DataFrame."""
        wb = _create_workbook(
            headers=["Col1", "Col2"],
            types=["texto", "entero"],
            examples=["abc", "123"],
            data_rows=[],
        )
        schema = _create_schema(["Col1", "Col2"], ["texto", "entero"], ["abc", "123"])

        df = converter.convert(wb, schema)

        assert isinstance(df, pd.DataFrame)
        assert list(df.columns) == ["Col1", "Col2"]
        assert len(df) == 0

    def test_skips_metadata_rows(self, converter):
        """Rows 2 and 3 (types and examples) are not included in the DataFrame."""
        wb = _create_workbook(
            headers=["Producto"],
            types=["texto"],
            examples=["Laptop"],
            data_rows=[["Teclado"], ["Mouse"]],
        )
        schema = _create_schema(["Producto"], ["texto"], ["Laptop"])

        df = converter.convert(wb, schema)

        # Should only have data rows, not metadata
        assert len(df) == 2
        assert "texto" not in df["Producto"].values
        assert "Laptop" not in df["Producto"].values
        assert "Teclado" in df["Producto"].values
        assert "Mouse" in df["Producto"].values


class TestDataFrameConverterSchemaFiltering:
    """Test that only schema columns are included."""

    def test_only_schema_columns_are_included(self, converter):
        """Columns not in the schema are excluded from the DataFrame."""
        # Workbook has 4 columns but schema only references columns 1 and 3
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="A")
        ws.cell(row=1, column=2, value="B")
        ws.cell(row=1, column=3, value="C")
        ws.cell(row=1, column=4, value="D")

        ws.cell(row=2, column=1, value="texto")
        ws.cell(row=2, column=2, value="texto")
        ws.cell(row=2, column=3, value="entero")
        ws.cell(row=2, column=4, value="texto")

        ws.cell(row=3, column=1, value="ex1")
        ws.cell(row=3, column=2, value="ex2")
        ws.cell(row=3, column=3, value="ex3")
        ws.cell(row=3, column=4, value="ex4")

        ws.cell(row=4, column=1, value="val1")
        ws.cell(row=4, column=2, value="val2")
        ws.cell(row=4, column=3, value=42)
        ws.cell(row=4, column=4, value="val4")

        # Schema only has columns 1 and 3
        schema = ColumnSchema(
            columns=[
                ColumnDef(index=1, name="A", data_type="texto", example_value="ex1"),
                ColumnDef(index=3, name="C", data_type="entero", example_value="ex3"),
            ]
        )

        df = converter.convert(wb, schema)

        assert list(df.columns) == ["A", "C"]
        assert len(df) == 1
        assert df.iloc[0]["A"] == "val1"
        assert df.iloc[0]["C"] == 42


class TestDataFrameConverterEdgeCases:
    """Test edge cases and special values."""

    def test_null_values_in_data(self, converter):
        """Cells with None values are preserved in the DataFrame as NaN."""
        wb = _create_workbook(
            headers=["X", "Y"],
            types=["texto", "entero"],
            examples=["a", "1"],
            data_rows=[["hello", None], [None, 99]],
        )
        schema = _create_schema(["X", "Y"], ["texto", "entero"], ["a", "1"])

        df = converter.convert(wb, schema)

        assert len(df) == 2
        assert df.iloc[0]["X"] == "hello"
        assert pd.isna(df.iloc[0]["Y"])
        assert pd.isna(df.iloc[1]["X"])
        assert df.iloc[1]["Y"] == 99

    def test_completely_empty_rows_are_filtered(self, converter):
        """Rows where all schema columns are None are excluded."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=2, column=1, value="texto")
        ws.cell(row=3, column=1, value="Juan")
        ws.cell(row=4, column=1, value="Carlos")
        # Row 5 is completely empty (no value set)
        ws.cell(row=6, column=1, value="Ana")

        schema = _create_schema(["Name"], ["texto"], ["Juan"])

        df = converter.convert(wb, schema)

        assert len(df) == 2
        assert df.iloc[0]["Name"] == "Carlos"
        assert df.iloc[1]["Name"] == "Ana"

    def test_special_characters_in_data(self, converter):
        """Data with special characters (accents, ñ) is preserved."""
        wb = _create_workbook(
            headers=["Descripción"],
            types=["texto"],
            examples=["Ejemplo con ñ"],
            data_rows=[["Piña colada"], ["Año nuevo"], ["Ñoño"]],
        )
        schema = _create_schema(["Descripción"], ["texto"], ["Ejemplo con ñ"])

        df = converter.convert(wb, schema)

        assert len(df) == 3
        assert df.iloc[0]["Descripción"] == "Piña colada"
        assert df.iloc[1]["Descripción"] == "Año nuevo"
        assert df.iloc[2]["Descripción"] == "Ñoño"

    def test_numeric_data_preserved(self, converter):
        """Numeric values retain their type in the DataFrame."""
        wb = _create_workbook(
            headers=["Precio", "Cantidad"],
            types=["decimal", "entero"],
            examples=["19.99", "5"],
            data_rows=[[29.99, 10], [5.50, 3]],
        )
        schema = _create_schema(
            ["Precio", "Cantidad"], ["decimal", "entero"], ["19.99", "5"]
        )

        df = converter.convert(wb, schema)

        assert len(df) == 2
        assert df.iloc[0]["Precio"] == 29.99
        assert df.iloc[0]["Cantidad"] == 10
        assert df.iloc[1]["Precio"] == 5.50
        assert df.iloc[1]["Cantidad"] == 3

    def test_max_columns_schema(self, converter):
        """Works correctly with the maximum allowed 8 columns."""
        headers = [f"Col{i}" for i in range(1, 9)]
        types = ["texto"] * 8
        examples = [f"ex{i}" for i in range(1, 9)]
        data_rows = [[f"val{i}" for i in range(1, 9)]]

        wb = _create_workbook(headers, types, examples, data_rows)
        schema = _create_schema(headers, types, examples)

        df = converter.convert(wb, schema)

        assert len(df.columns) == 8
        assert list(df.columns) == headers
        assert len(df) == 1
