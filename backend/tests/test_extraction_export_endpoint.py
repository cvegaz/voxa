"""Unit tests for GET /api/extraction/export/{session_id} endpoint."""

from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@pytest.fixture
def mock_pool():
    return MagicMock()


@pytest.fixture
def client(mock_pool):
    app.state.pool = mock_pool
    return TestClient(app, raise_server_exceptions=False)


def _schema_json():
    return {
        "columns": [
            {"index": 1, "name": "Equipo_Local", "data_type": "texto", "example_value": "Borregos"},
            {"index": 2, "name": "Estadio", "data_type": "texto", "example_value": "Disney"},
        ]
    }


class TestExportExcel:
    """Tests for the on-demand Excel download."""

    @patch("app.routes.extraction_routes.ExtractionRepository")
    def test_export_returns_xlsx_with_header_and_rows(self, MockRepo, client):
        """Returns a valid .xlsx: header row + one data row, correct headers."""
        session_id = str(uuid4())

        mock_repo = MockRepo.return_value
        mock_repo.get_export_context = AsyncMock(
            return_value={"schema_json": _schema_json(), "file_name": "partido.xlsx"}
        )
        mock_repo.get_records = AsyncMock(
            return_value=[
                {"record_json": {"Equipo_Local": "Pumas", "Estadio": "Azteca"}},
            ]
        )

        response = client.get(f"/api/extraction/export/{session_id}")

        assert response.status_code == 200
        assert response.headers["content-type"] == XLSX_MEDIA_TYPE
        assert 'attachment; filename="partido.xlsx"' in response.headers["content-disposition"]

        sheet = load_workbook(BytesIO(response.content)).active
        assert [c.value for c in sheet[1]] == ["Equipo_Local", "Estadio"]
        assert [c.value for c in sheet[2]] == ["Pumas", "Azteca"]
        assert sheet.max_row == 2

    @patch("app.routes.extraction_routes.ExtractionRepository")
    def test_export_empty_session_has_header_only(self, MockRepo, client):
        """A session with no records exports just the header row."""
        session_id = str(uuid4())

        mock_repo = MockRepo.return_value
        mock_repo.get_export_context = AsyncMock(
            return_value={"schema_json": _schema_json(), "file_name": "partido.xlsx"}
        )
        mock_repo.get_records = AsyncMock(return_value=[])

        response = client.get(f"/api/extraction/export/{session_id}")

        assert response.status_code == 200
        sheet = load_workbook(BytesIO(response.content)).active
        assert sheet.max_row == 1

    @patch("app.routes.extraction_routes.ExtractionRepository")
    def test_export_missing_session_returns_404(self, MockRepo, client):
        """Unknown session returns 404 SESSION_NOT_FOUND."""
        session_id = str(uuid4())

        mock_repo = MockRepo.return_value
        mock_repo.get_export_context = AsyncMock(return_value=None)

        response = client.get(f"/api/extraction/export/{session_id}")

        assert response.status_code == 404
        assert response.json()["errorCode"] == "SESSION_NOT_FOUND"

    @patch("app.routes.extraction_routes.ExtractionRepository")
    def test_export_falls_back_to_default_filename(self, MockRepo, client):
        """A session without a stored file name still downloads with a default."""
        session_id = str(uuid4())

        mock_repo = MockRepo.return_value
        mock_repo.get_export_context = AsyncMock(
            return_value={"schema_json": _schema_json(), "file_name": None}
        )
        mock_repo.get_records = AsyncMock(return_value=[])

        response = client.get(f"/api/extraction/export/{session_id}")

        assert response.status_code == 200
        assert 'filename="export.xlsx"' in response.headers["content-disposition"]
