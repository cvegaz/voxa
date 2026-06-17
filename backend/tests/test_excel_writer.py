"""Unit tests for the ExcelWriter service."""

import os
import tempfile

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from app.models import ColumnDef, ColumnSchema
from app.services.excel_writer import ExcelWriter


@pytest.fixture
def schema():
    """A simple 3-column schema."""
    return ColumnSchema(
        columns=[
            ColumnDef(index=1, name="Nombre", data_type="texto", example_value="Juan"),
            ColumnDef(index=2, name="Edad", data_type="número entero", example_value="30"),
            ColumnDef(index=3, name="Ciudad", data_type="texto", example_value="Madrid"),
        ]
    )


@pytest.fixture
def excel_file(schema):
    """Create a temporary .xlsx file with header rows 1-3."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()

    wb = Workbook()
    ws = wb.active

    # Row 1: column names
    for col_def in schema.columns:
        ws.cell(row=1, column=col_def.index, value=col_def.name)

    # Row 2: data types
    for col_def in schema.columns:
        ws.cell(row=2, column=col_def.index, value=col_def.data_type)

    # Row 3: example values
    for col_def in schema.columns:
        ws.cell(row=3, column=col_def.index, value=col_def.example_value)

    wb.save(tmp.name)
    yield tmp.name
    os.unlink(tmp.name)


@pytest.fixture
def writer():
    return ExcelWriter()


class TestExcelWriterWriteNewData:
    """Test writing new data rows to a file with only headers."""

    def test_writes_single_row(self, writer, excel_file, schema):
        df = pd.DataFrame([{"Nombre": "Ana", "Edad": "25", "Ciudad": "Barcelona"}])

        writer.write(df, excel_file, schema)

        wb = load_workbook(excel_file)
        ws = wb.active
        assert ws.cell(row=4, column=1).value == "Ana"
        assert ws.cell(row=4, column=2).value == "25"
        assert ws.cell(row=4, column=3).value == "Barcelona"

    def test_writes_multiple_rows(self, writer, excel_file, schema):
        df = pd.DataFrame(
            [
                {"Nombre": "Ana", "Edad": "25", "Ciudad": "Barcelona"},
                {"Nombre": "Luis", "Edad": "40", "Ciudad": "Sevilla"},
            ]
        )

        writer.write(df, excel_file, schema)

        wb = load_workbook(excel_file)
        ws = wb.active
        assert ws.cell(row=4, column=1).value == "Ana"
        assert ws.cell(row=5, column=1).value == "Luis"
        assert ws.cell(row=5, column=2).value == "40"
        assert ws.cell(row=5, column=3).value == "Sevilla"


class TestExcelWriterOverwriteExistingData:
    """Test that existing data rows are replaced."""

    def test_overwrites_old_data(self, writer, excel_file, schema):
        # Write initial data
        df_initial = pd.DataFrame(
            [{"Nombre": "Old1", "Edad": "10", "Ciudad": "OldCity"}]
        )
        writer.write(df_initial, excel_file, schema)

        # Write new data (should replace old)
        df_new = pd.DataFrame(
            [
                {"Nombre": "New1", "Edad": "20", "Ciudad": "NewCity"},
                {"Nombre": "New2", "Edad": "30", "Ciudad": "NewCity2"},
            ]
        )
        writer.write(df_new, excel_file, schema)

        wb = load_workbook(excel_file)
        ws = wb.active
        assert ws.cell(row=4, column=1).value == "New1"
        assert ws.cell(row=5, column=1).value == "New2"
        # Old data should be gone (no row 6 with data)
        assert ws.cell(row=6, column=1).value is None


class TestExcelWriterPreservesHeaders:
    """Test that header rows (1-3) are preserved after write."""

    def test_headers_unchanged_after_write(self, writer, excel_file, schema):
        df = pd.DataFrame([{"Nombre": "Test", "Edad": "99", "Ciudad": "Test City"}])

        writer.write(df, excel_file, schema)

        wb = load_workbook(excel_file)
        ws = wb.active
        # Row 1: column names
        assert ws.cell(row=1, column=1).value == "Nombre"
        assert ws.cell(row=1, column=2).value == "Edad"
        assert ws.cell(row=1, column=3).value == "Ciudad"
        # Row 2: data types
        assert ws.cell(row=2, column=1).value == "texto"
        assert ws.cell(row=2, column=2).value == "número entero"
        assert ws.cell(row=2, column=3).value == "texto"
        # Row 3: example values
        assert ws.cell(row=3, column=1).value == "Juan"
        assert ws.cell(row=3, column=2).value == "30"
        assert ws.cell(row=3, column=3).value == "Madrid"


class TestExcelWriterEmptyDataFrame:
    """Test writing an empty DataFrame clears existing data."""

    def test_empty_dataframe_clears_data(self, writer, excel_file, schema):
        # Write some data first
        df_initial = pd.DataFrame(
            [{"Nombre": "Data", "Edad": "50", "Ciudad": "City"}]
        )
        writer.write(df_initial, excel_file, schema)

        # Write empty DataFrame
        df_empty = pd.DataFrame(columns=["Nombre", "Edad", "Ciudad"])
        writer.write(df_empty, excel_file, schema)

        wb = load_workbook(excel_file)
        ws = wb.active
        # Row 4 should be empty
        assert ws.cell(row=4, column=1).value is None
        # Headers still preserved
        assert ws.cell(row=1, column=1).value == "Nombre"


class TestExcelWriterFileNotFound:
    """Test that FileNotFoundError is raised for non-existent file."""

    def test_raises_file_not_found(self, writer, schema):
        df = pd.DataFrame([{"Nombre": "Test", "Edad": "1", "Ciudad": "X"}])

        with pytest.raises(FileNotFoundError):
            writer.write(df, "/nonexistent/path/file.xlsx", schema)
