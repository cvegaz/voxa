"""Unit tests for SchemaExtractor service."""

import pytest
from openpyxl import Workbook

from app.services.schema_extractor import SchemaExtractor


def _create_valid_workbook(num_columns: int = 3) -> Workbook:
    """Create a valid workbook with the given number of columns."""
    wb = Workbook()
    ws = wb.active
    for col in range(1, num_columns + 1):
        ws.cell(row=1, column=col, value=f"Columna_{col}")
        ws.cell(row=2, column=col, value="texto")
        ws.cell(row=3, column=col, value=f"ejemplo_{col}")
    return wb


class TestSchemaExtractorBasic:
    """Tests for basic schema extraction functionality."""

    def setup_method(self):
        self.extractor = SchemaExtractor()

    def test_extracts_single_column(self):
        wb = _create_valid_workbook(1)
        schema = self.extractor.extract(wb)
        assert len(schema.columns) == 1
        col = schema.columns[0]
        assert col.index == 1
        assert col.name == "Columna_1"
        assert col.data_type == "texto"
        assert col.example_value == "ejemplo_1"

    def test_extracts_multiple_columns(self):
        wb = _create_valid_workbook(4)
        schema = self.extractor.extract(wb)
        assert len(schema.columns) == 4
        for i, col in enumerate(schema.columns, start=1):
            assert col.index == i
            assert col.name == f"Columna_{i}"
            assert col.data_type == "texto"
            assert col.example_value == f"ejemplo_{i}"

    def test_extracts_max_eight_columns(self):
        wb = _create_valid_workbook(8)
        schema = self.extractor.extract(wb)
        assert len(schema.columns) == 8
        assert schema.columns[7].index == 8

    def test_skips_columns_with_empty_name(self):
        wb = Workbook()
        ws = wb.active
        # Column 1: valid
        ws.cell(row=1, column=1, value="Nombre")
        ws.cell(row=2, column=1, value="texto")
        ws.cell(row=3, column=1, value="Juan")
        # Column 2: empty name (should be skipped)
        ws.cell(row=1, column=2, value="")
        ws.cell(row=2, column=2, value="número")
        ws.cell(row=3, column=2, value="42")
        # Column 3: valid
        ws.cell(row=1, column=3, value="Edad")
        ws.cell(row=2, column=3, value="número entero")
        ws.cell(row=3, column=3, value="25")

        schema = self.extractor.extract(wb)
        assert len(schema.columns) == 2
        assert schema.columns[0].name == "Nombre"
        assert schema.columns[0].index == 1
        assert schema.columns[1].name == "Edad"
        assert schema.columns[1].index == 3

    def test_skips_columns_with_none_name(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Nombre")
        ws.cell(row=2, column=1, value="texto")
        ws.cell(row=3, column=1, value="Juan")
        # Column 2: None name (default empty cell)
        ws.cell(row=2, column=2, value="número")
        ws.cell(row=3, column=2, value="42")

        schema = self.extractor.extract(wb)
        assert len(schema.columns) == 1
        assert schema.columns[0].name == "Nombre"

    def test_skips_columns_with_whitespace_only_name(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Nombre")
        ws.cell(row=2, column=1, value="texto")
        ws.cell(row=3, column=1, value="Juan")
        ws.cell(row=1, column=2, value="   ")
        ws.cell(row=2, column=2, value="texto")
        ws.cell(row=3, column=2, value="dato")

        schema = self.extractor.extract(wb)
        assert len(schema.columns) == 1
        assert schema.columns[0].name == "Nombre"


class TestSchemaExtractorValueConversion:
    """Tests for cell value to string conversion."""

    def setup_method(self):
        self.extractor = SchemaExtractor()

    def test_converts_numeric_values_to_string(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Edad")
        ws.cell(row=2, column=1, value="número entero")
        ws.cell(row=3, column=1, value=25)

        schema = self.extractor.extract(wb)
        assert schema.columns[0].example_value == "25"

    def test_converts_numeric_name_to_string(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=123)
        ws.cell(row=2, column=1, value="id")
        ws.cell(row=3, column=1, value=456)

        schema = self.extractor.extract(wb)
        assert schema.columns[0].name == "123"
        assert schema.columns[0].data_type == "id"
        assert schema.columns[0].example_value == "456"

    def test_converts_float_values_to_string(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Precio")
        ws.cell(row=2, column=1, value="decimal")
        ws.cell(row=3, column=1, value=19.99)

        schema = self.extractor.extract(wb)
        assert schema.columns[0].example_value == "19.99"

    def test_strips_whitespace_from_values(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="  Nombre  ")
        ws.cell(row=2, column=1, value="  texto  ")
        ws.cell(row=3, column=1, value="  Juan  ")

        schema = self.extractor.extract(wb)
        assert schema.columns[0].name == "Nombre"
        assert schema.columns[0].data_type == "texto"
        assert schema.columns[0].example_value == "Juan"


class TestSchemaExtractorSpecialCharacters:
    """Tests for special characters handling."""

    def setup_method(self):
        self.extractor = SchemaExtractor()

    def test_handles_spanish_characters(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Dirección")
        ws.cell(row=1, column=2, value="Año")
        ws.cell(row=2, column=1, value="texto")
        ws.cell(row=2, column=2, value="número entero")
        ws.cell(row=3, column=1, value="Calle Ñoño #123")
        ws.cell(row=3, column=2, value="1990")

        schema = self.extractor.extract(wb)
        assert len(schema.columns) == 2
        assert schema.columns[0].name == "Dirección"
        assert schema.columns[0].example_value == "Calle Ñoño #123"
        assert schema.columns[1].name == "Año"

    def test_handles_symbols_in_values(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Email")
        ws.cell(row=2, column=1, value="correo electrónico")
        ws.cell(row=3, column=1, value="user@example.com")

        schema = self.extractor.extract(wb)
        assert schema.columns[0].example_value == "user@example.com"


class TestSchemaExtractorColumnOrder:
    """Tests verifying column order preservation."""

    def setup_method(self):
        self.extractor = SchemaExtractor()

    def test_preserves_column_order(self):
        wb = Workbook()
        ws = wb.active
        names = ["Zebra", "Apple", "Mango"]
        for i, name in enumerate(names, start=1):
            ws.cell(row=1, column=i, value=name)
            ws.cell(row=2, column=i, value="texto")
            ws.cell(row=3, column=i, value=f"ej_{name}")

        schema = self.extractor.extract(wb)
        assert [c.name for c in schema.columns] == ["Zebra", "Apple", "Mango"]

    def test_index_reflects_original_column_position(self):
        """When columns are skipped, index reflects original Excel column."""
        wb = Workbook()
        ws = wb.active
        # Col 1: valid
        ws.cell(row=1, column=1, value="A")
        ws.cell(row=2, column=1, value="tipo_a")
        ws.cell(row=3, column=1, value="ej_a")
        # Col 2: empty name (skipped)
        ws.cell(row=1, column=2, value="")
        ws.cell(row=2, column=2, value="tipo_b")
        ws.cell(row=3, column=2, value="ej_b")
        # Col 3: valid
        ws.cell(row=1, column=3, value="C")
        ws.cell(row=2, column=3, value="tipo_c")
        ws.cell(row=3, column=3, value="ej_c")

        schema = self.extractor.extract(wb)
        assert schema.columns[0].index == 1
        assert schema.columns[1].index == 3
