"""Unit and integration tests for POST /api/templates/upload endpoint."""

import io
import json
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app


def _create_valid_xlsx(num_columns: int = 3, num_data_rows: int = 2) -> bytes:
    """Create a valid Excel file in memory and return its bytes."""
    wb = Workbook()
    ws = wb.active
    for col in range(1, num_columns + 1):
        ws.cell(row=1, column=col, value=f"Columna_{col}")
        ws.cell(row=2, column=col, value="texto")
        ws.cell(row=3, column=col, value=f"ejemplo_{col}")
    # Add some data rows
    for row in range(4, 4 + num_data_rows):
        for col in range(1, num_columns + 1):
            ws.cell(row=row, column=col, value=f"dato_{row}_{col}")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _create_invalid_extension_file() -> bytes:
    """Create content for a non-xlsx file."""
    return b"This is a plain text file"


def _create_too_many_columns_xlsx() -> bytes:
    """Create an Excel file with more than 8 columns."""
    wb = Workbook()
    ws = wb.active
    for col in range(1, 11):  # 10 columns
        ws.cell(row=1, column=col, value=f"Col_{col}")
        ws.cell(row=2, column=col, value="texto")
        ws.cell(row=3, column=col, value=f"ej_{col}")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _create_missing_data_types_xlsx() -> bytes:
    """Create an Excel file with missing data types in row 2."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Nombre")
    ws.cell(row=1, column=2, value="Edad")
    ws.cell(row=2, column=1, value="texto")
    # Column 2 row 2 is empty
    ws.cell(row=3, column=1, value="Juan")
    ws.cell(row=3, column=2, value="25")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _create_empty_header_xlsx() -> bytes:
    """Create an Excel file with empty header row."""
    wb = Workbook()
    ws = wb.active
    # Row 1 is empty, add content in row 2
    ws.cell(row=2, column=1, value="texto")
    ws.cell(row=3, column=1, value="ejemplo")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@pytest.fixture
def mock_pool():
    """Create a mock asyncpg pool."""
    pool = MagicMock()
    return pool


@pytest.fixture
def client(mock_pool):
    """Create a test client with mocked database pool."""
    app.state.pool = mock_pool
    return TestClient(app, raise_server_exceptions=False)


class TestUploadEndpointSuccess:
    """Tests for successful file upload."""

    def test_valid_upload_returns_200(self, client, mock_pool):
        """Valid .xlsx file returns 200 with session_id, schema, and file_name."""
        session_id = str(uuid4())

        # Mock the pool connection for replace_previous_sessions and create_session
        mock_conn = AsyncMock()
        mock_conn.execute = AsyncMock(return_value="UPDATE 0")
        mock_conn.fetchrow = AsyncMock(return_value={"id": session_id})

        mock_pool.acquire = MagicMock()
        mock_pool.acquire.return_value.__aenter__ = AsyncMock(return_value=mock_conn)
        mock_pool.acquire.return_value.__aexit__ = AsyncMock(return_value=None)

        file_content = _create_valid_xlsx(num_columns=3)
        response = client.post(
            "/api/templates/upload",
            files={"file": ("plantilla.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["session_id"] == session_id
        assert data["file_name"] == "plantilla.xlsx"
        assert "schema" in data
        assert len(data["schema"]["columns"]) == 3

    def test_valid_upload_schema_structure(self, client, mock_pool):
        """Uploaded schema contains correct column definitions."""
        session_id = str(uuid4())

        mock_conn = AsyncMock()
        mock_conn.execute = AsyncMock(return_value="UPDATE 0")
        mock_conn.fetchrow = AsyncMock(return_value={"id": session_id})

        mock_pool.acquire = MagicMock()
        mock_pool.acquire.return_value.__aenter__ = AsyncMock(return_value=mock_conn)
        mock_pool.acquire.return_value.__aexit__ = AsyncMock(return_value=None)

        file_content = _create_valid_xlsx(num_columns=2)
        response = client.post(
            "/api/templates/upload",
            files={"file": ("test.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        schema = response.json()["schema"]
        columns = schema["columns"]

        assert columns[0]["index"] == 1
        assert columns[0]["name"] == "Columna_1"
        assert columns[0]["data_type"] == "texto"
        assert columns[0]["example_value"] == "ejemplo_1"

        assert columns[1]["index"] == 2
        assert columns[1]["name"] == "Columna_2"
        assert columns[1]["data_type"] == "texto"
        assert columns[1]["example_value"] == "ejemplo_2"

    def test_valid_upload_calls_replace_previous_sessions(self, client, mock_pool):
        """Upload replaces previous sessions before creating a new one."""
        session_id = str(uuid4())

        mock_conn = AsyncMock()
        mock_conn.execute = AsyncMock(return_value="UPDATE 2")
        mock_conn.fetchrow = AsyncMock(return_value={"id": session_id})

        mock_pool.acquire = MagicMock()
        mock_pool.acquire.return_value.__aenter__ = AsyncMock(return_value=mock_conn)
        mock_pool.acquire.return_value.__aexit__ = AsyncMock(return_value=None)

        file_content = _create_valid_xlsx()
        response = client.post(
            "/api/templates/upload",
            files={"file": ("test.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        # Verify execute was called (for replace_previous_sessions)
        assert mock_conn.execute.called

    def test_upload_with_8_columns(self, client, mock_pool):
        """File with exactly 8 columns (max allowed) succeeds."""
        session_id = str(uuid4())

        mock_conn = AsyncMock()
        mock_conn.execute = AsyncMock(return_value="UPDATE 0")
        mock_conn.fetchrow = AsyncMock(return_value={"id": session_id})

        mock_pool.acquire = MagicMock()
        mock_pool.acquire.return_value.__aenter__ = AsyncMock(return_value=mock_conn)
        mock_pool.acquire.return_value.__aexit__ = AsyncMock(return_value=None)

        file_content = _create_valid_xlsx(num_columns=8)
        response = client.post(
            "/api/templates/upload",
            files={"file": ("max_cols.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        assert len(response.json()["schema"]["columns"]) == 8


class TestUploadEndpointValidationErrors:
    """Tests for validation failure cases returning 422."""

    def test_invalid_extension_returns_422(self, client):
        """Non-.xlsx file returns 422 with INVALID_EXTENSION error."""
        response = client.post(
            "/api/templates/upload",
            files={"file": ("data.csv", _create_invalid_extension_file(), "text/csv")},
        )

        assert response.status_code == 422
        data = response.json()["detail"]
        assert data["error_code"] == "INVALID_EXTENSION"

    def test_too_many_columns_returns_422(self, client):
        """File with >8 columns returns 422 with TOO_MANY_COLUMNS error."""
        file_content = _create_too_many_columns_xlsx()
        response = client.post(
            "/api/templates/upload",
            files={"file": ("many_cols.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 422
        data = response.json()["detail"]
        assert data["error_code"] == "TOO_MANY_COLUMNS"

    def test_missing_data_types_returns_422(self, client):
        """File with incomplete row 2 returns 422 with MISSING_DATA_TYPES error."""
        file_content = _create_missing_data_types_xlsx()
        response = client.post(
            "/api/templates/upload",
            files={"file": ("missing_types.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 422
        data = response.json()["detail"]
        assert data["error_code"] == "MISSING_DATA_TYPES"

    def test_empty_header_returns_422(self, client):
        """File with empty row 1 returns 422 with EMPTY_HEADER_ROW error."""
        file_content = _create_empty_header_xlsx()
        response = client.post(
            "/api/templates/upload",
            files={"file": ("no_headers.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 422
        data = response.json()["detail"]
        assert data["error_code"] == "EMPTY_HEADER_ROW"

    def test_corrupt_file_returns_422(self, client):
        """Corrupt/unreadable file returns 422 with UNREADABLE_FILE error."""
        response = client.post(
            "/api/templates/upload",
            files={"file": ("bad.xlsx", b"not excel content", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 422
        data = response.json()["detail"]
        assert data["error_code"] == "UNREADABLE_FILE"

    def test_no_file_field_returns_422(self, client):
        """Request without file field returns 422."""
        response = client.post("/api/templates/upload")
        assert response.status_code == 422


class TestUploadEndpointResponseFormat:
    """Tests for response format and structure."""

    def test_response_contains_required_fields(self, client, mock_pool):
        """Response includes session_id, schema, and file_name."""
        session_id = str(uuid4())

        mock_conn = AsyncMock()
        mock_conn.execute = AsyncMock(return_value="UPDATE 0")
        mock_conn.fetchrow = AsyncMock(return_value={"id": session_id})

        mock_pool.acquire = MagicMock()
        mock_pool.acquire.return_value.__aenter__ = AsyncMock(return_value=mock_conn)
        mock_pool.acquire.return_value.__aexit__ = AsyncMock(return_value=None)

        file_content = _create_valid_xlsx()
        response = client.post(
            "/api/templates/upload",
            files={"file": ("my_template.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        data = response.json()
        assert "session_id" in data
        assert "schema" in data
        assert "file_name" in data
        assert data["file_name"] == "my_template.xlsx"

    def test_schema_columns_have_all_fields(self, client, mock_pool):
        """Each column in schema has index, name, data_type, example_value."""
        session_id = str(uuid4())

        mock_conn = AsyncMock()
        mock_conn.execute = AsyncMock(return_value="UPDATE 0")
        mock_conn.fetchrow = AsyncMock(return_value={"id": session_id})

        mock_pool.acquire = MagicMock()
        mock_pool.acquire.return_value.__aenter__ = AsyncMock(return_value=mock_conn)
        mock_pool.acquire.return_value.__aexit__ = AsyncMock(return_value=None)

        file_content = _create_valid_xlsx(num_columns=2)
        response = client.post(
            "/api/templates/upload",
            files={"file": ("test.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        columns = response.json()["schema"]["columns"]
        for col in columns:
            assert "index" in col
            assert "name" in col
            assert "data_type" in col
            assert "example_value" in col
